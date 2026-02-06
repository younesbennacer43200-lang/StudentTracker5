# main_android.py - Android-Optimized Student Tracker (NO PANDAS)
# Programmed by: Younes Bennacer  
# Version: 2.0.1 Android Edition
# Lightweight version for Android APK - Works reliably!

import os
import sqlite3
import json
from datetime import datetime
from collections import defaultdict
import logging
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.spinner import Spinner
from kivy.uix.progressbar import ProgressBar
from kivy.uix.checkbox import CheckBox
from kivy.core.window import Window
from kivy.metrics import dp, sp
from kivy.graphics import Color, RoundedRectangle, Line
from kivy.clock import Clock
from kivy.properties import StringProperty
import threading

# ============================================
# CONFIGURATION
# ============================================
class Config:
    APP_NAME = "Student Tracker Pro"
    DEVELOPER = "Younes Bennacer"
    VERSION = "2.0.1 Android"
    DB_NAME = 'student_tracker.db'
    AUTO_BACKUP_INTERVAL = 3600
    MAX_STUDENTS_PER_GROUP = 60
    MAX_GROUPS = 30
    SEMESTER_SECTIONS = 15
    TOTAL_GROUPS_LOADED = 20

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================
# COLORS
# ============================================
PRIMARY_COLOR = (0.2, 0.6, 0.86, 1)
SUCCESS_COLOR = (0.3, 0.7, 0.4, 1)
WARNING_COLOR = (1, 0.6, 0.2, 1)
ERROR_COLOR = (0.9, 0.3, 0.3, 1)
BACKGROUND_COLOR = (0.95, 0.95, 0.97, 1)
CARD_COLOR = (1, 1, 1, 1)
TEXT_COLOR = (0.2, 0.2, 0.2, 1)
BORDER_COLOR = (0.85, 0.85, 0.87, 1)

# ============================================
# ANDROID SUPPORT
# ============================================
from kivy.utils import platform

if platform == 'android':
    from android.permissions import request_permissions, Permission, check_permission
    from android.storage import primary_external_storage_path
    
    def request_android_permissions():
        try:
            perms = [Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE]
            request_permissions(perms)
            logger.info("Android permissions requested")
        except Exception as e:
            logger.error(f"Permission error: {e}")
    
    def get_storage_path():
        try:
            return primary_external_storage_path()
        except:
            return '/sdcard'
else:
    def request_android_permissions():
        pass
    
    def get_storage_path():
        return os.path.expanduser('~')

# ============================================
# EXCEL EXPORT WITHOUT PANDAS (Pure Python + openpyxl)
# ============================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")

# ============================================
# DATABASE MANAGER
# ============================================
class StudentTrackerDB:
    """Lightweight database manager for Android"""
    
    def __init__(self, db_name=Config.DB_NAME):
        self.db_name = db_name
        self.lock = threading.Lock()
        self.init_database()
        logger.info(f"Database initialized: {db_name}")
    
    def get_connection(self):
        conn = sqlite3.connect(self.db_name, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn
    
    def init_database(self):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    matricule TEXT UNIQUE NOT NULL,
                    nom TEXT NOT NULL,
                    prenom TEXT NOT NULL,
                    groupe TEXT NOT NULL,
                    date_added TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS grades (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id INTEGER NOT NULL,
                    section_number INTEGER CHECK(section_number >= 1 AND section_number <= 15),
                    note REAL,
                    absent INTEGER DEFAULT 0,
                    justifie INTEGER DEFAULT 0,
                    observation TEXT,
                    date_recorded TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
                    UNIQUE(student_id, section_number)
                )
            ''')
            
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_students_groupe ON students(groupe)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_grades_student ON grades(student_id, section_number)')
            
            conn.commit()
            conn.close()
    
    def add_student(self, matricule, nom, prenom, groupe):
        with self.lock:
            try:
                conn = self.get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    'INSERT INTO students (matricule, nom, prenom, groupe) VALUES (?, ?, ?, ?)',
                    (matricule, nom, prenom, groupe)
                )
                conn.commit()
                student_id = cursor.lastrowid
                conn.close()
                return True, "Student added", student_id
            except sqlite3.IntegrityError:
                return False, "Duplicate matricule", None
            except Exception as e:
                return False, str(e), None
    
    def add_or_update_grade(self, student_id, section, note=None, absent=0, justifie=0, obs=''):
        with self.lock:
            try:
                conn = self.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO grades (student_id, section_number, note, absent, justifie, observation)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(student_id, section_number)
                    DO UPDATE SET note=?, absent=?, justifie=?, observation=?
                ''', (student_id, section, note, absent, justifie, obs, note, absent, justifie, obs))
                conn.commit()
                conn.close()
                return True, "Grade saved"
            except Exception as e:
                return False, str(e)
    
    def get_student_grades(self, student_id):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT section_number, note, absent, justifie, observation
                FROM grades WHERE student_id = ? ORDER BY section_number
            ''', (student_id,))
            grades = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return grades
    
    def calculate_average(self, student_id):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT AVG(note) as avg, COUNT(note) as cnt
                FROM grades WHERE student_id = ? AND note IS NOT NULL
            ''', (student_id,))
            result = cursor.fetchone()
            conn.close()
            if result and result['cnt'] > 0:
                return round(result['avg'], 2), result['cnt']
            return None, 0
    
    def get_all_students(self, groupe=None):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            if groupe:
                cursor.execute('SELECT * FROM students WHERE groupe = ? ORDER BY nom', (groupe,))
            else:
                cursor.execute('SELECT * FROM students ORDER BY groupe, nom')
            students = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return students
    
    def get_all_groups(self):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT groupe FROM students ORDER BY groupe')
            groups = [row[0] for row in cursor.fetchall()]
            conn.close()
            return groups
    
    def import_students_from_roster(self, groupe, students_list):
        with self.lock:
            success = 0
            errors = 0
            conn = self.get_connection()
            cursor = conn.cursor()
            
            for student in students_list:
                try:
                    cursor.execute(
                        'INSERT INTO students (matricule, nom, prenom, groupe) VALUES (?, ?, ?, ?)',
                        (student['matricule'], student['nom'], student['prenom'], groupe)
                    )
                    success += 1
                except:
                    errors += 1
            
            conn.commit()
            conn.close()
            return success, errors, []
    
    def export_to_excel(self, output_path, groupe):
        """Export using pure openpyxl (no pandas)"""
        if not EXCEL_AVAILABLE:
            return False, "Excel library not available"
        
        try:
            students = self.get_all_students(groupe)
            if not students:
                return False, "No students found"
            
            wb = Workbook()
            ws = wb.active
            ws.title = groupe
            
            # Headers
            headers = ['Matricule', 'Nom', 'Prénom', 'Groupe']
            for i in range(1, Config.SEMESTER_SECTIONS + 1):
                headers.append(f'Section_{i}')
                headers.append(f'Absent_{i}')
            headers.extend(['Moyenne', 'Notes_Comptées'])
            
            ws.append(headers)
            
            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')
            
            # Data rows
            for student in students:
                row = [student['matricule'], student['nom'], student['prenom'], student['groupe']]
                
                grades = self.get_student_grades(student['id'])
                grade_dict = {g['section_number']: g for g in grades}
                
                for section in range(1, Config.SEMESTER_SECTIONS + 1):
                    if section in grade_dict:
                        row.append(grade_dict[section]['note'])
                        row.append('Oui' if grade_dict[section]['absent'] else 'Non')
                    else:
                        row.append(None)
                        row.append('Non')
                
                avg, count = self.calculate_average(student['id'])
                row.append(avg)
                row.append(count)
                
                ws.append(row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            wb.save(output_path)
            return True, f"Exported {len(students)} students"
        except Exception as e:
            logger.error(f"Export error: {e}")
            return False, str(e)
    
    def search_students(self, query, groupe=None):
        with self.lock:
            conn = self.get_connection()
            cursor = conn.cursor()
            pattern = f'%{query}%'
            if groupe:
                cursor.execute('''
                    SELECT * FROM students 
                    WHERE groupe = ? AND (nom LIKE ? OR prenom LIKE ? OR matricule LIKE ?)
                ''', (groupe, pattern, pattern, pattern))
            else:
                cursor.execute('''
                    SELECT * FROM students 
                    WHERE nom LIKE ? OR prenom LIKE ? OR matricule LIKE ?
                ''', (pattern, pattern, pattern))
            students = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return students
    
    def delete_student(self, student_id):
        with self.lock:
            try:
                conn = self.get_connection()
                cursor = conn.cursor()
                cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))
                conn.commit()
                conn.close()
                return True, "Deleted"
            except Exception as e:
                return False, str(e)
    
    def backup_database(self):
        try:
            import shutil
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = f'backup_{timestamp}.db'
            shutil.copy2(self.db_name, backup_path)
            return True, backup_path
        except Exception as e:
            return False, str(e)

# ============================================
# UI COMPONENTS
# ============================================
def show_popup(message, title='Info'):
    content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
    content.add_widget(Label(text=message, size_hint_y=0.7))
    btn = Button(text='OK', size_hint_y=None, height=dp(45))
    content.add_widget(btn)
    popup = Popup(title=title, content=content, size_hint=(0.8, 0.4))
    btn.bind(on_press=popup.dismiss)
    popup.open()

class ModernButton(Button):
    def __init__(self, color=PRIMARY_COLOR, **kwargs):
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0)
        self.btn_color = color
        self.color = (1, 1, 1, 1)
        self.font_size = sp(15)
        self.bind(pos=self.draw, size=self.draw)
    
    def draw(self, *args):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.btn_color)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(6)])

# ============================================
# MAIN SCREEN
# ============================================
class MainScreen(Screen):
    def __init__(self, db, **kwargs):
        super().__init__(**kwargs)
        self.db = db
        self.selected_groupe = None
        self.selected_student = None
        self.current_section = 1
        self.rosters = self.load_rosters()
        self.build_ui()
    
    def load_rosters(self):
        try:
            roster_file = 'student_rosters.json'
            if os.path.exists(roster_file):
                with open(roster_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Roster load error: {e}")
        return {}
    
    def build_ui(self):
        layout = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(10))
        
        # Header
        header = Label(
            text=f'{Config.APP_NAME} v{Config.VERSION}',
            size_hint_y=None,
            height=dp(50),
            font_size=sp(20),
            bold=True
        )
        layout.add_widget(header)
        
        # Controls
        controls = BoxLayout(size_hint_y=None, height=dp(100), orientation='vertical', spacing=dp(5))
        
        # Row 1: Group selection
        row1 = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(5))
        row1.add_widget(Label(text='Group:', size_hint_x=0.2))
        self.group_spinner = Spinner(text='Select', values=[], size_hint_x=0.4)
        self.group_spinner.bind(text=self.on_group_selected)
        row1.add_widget(self.group_spinner)
        
        load_btn = ModernButton(text='Load Roster', size_hint_x=0.4)
        load_btn.bind(on_press=self.load_roster)
        row1.add_widget(load_btn)
        controls.add_widget(row1)
        
        # Row 2: Section + Actions
        row2 = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(5))
        row2.add_widget(Label(text='Section:', size_hint_x=0.2))
        
        sections = [f'Section {i}' for i in range(1, Config.SEMESTER_SECTIONS + 1)]
        self.section_spinner = Spinner(text='Section 1', values=sections, size_hint_x=0.3)
        self.section_spinner.bind(text=self.on_section_change)
        row2.add_widget(self.section_spinner)
        
        export_btn = ModernButton(text='Export', color=SUCCESS_COLOR, size_hint_x=0.25)
        export_btn.bind(on_press=self.export_grades)
        row2.add_widget(export_btn)
        
        add_btn = ModernButton(text='Add', color=WARNING_COLOR, size_hint_x=0.25)
        add_btn.bind(on_press=self.add_student)
        row2.add_widget(add_btn)
        controls.add_widget(row2)
        
        layout.add_widget(controls)
        
        # Main area
        main = BoxLayout(spacing=dp(10))
        
        # Student list
        list_box = BoxLayout(orientation='vertical', size_hint_x=0.4)
        list_box.add_widget(Label(text='Students', size_hint_y=None, height=dp(30), bold=True))
        
        scroll = ScrollView()
        self.student_list = GridLayout(cols=1, spacing=dp(3), size_hint_y=None)
        self.student_list.bind(minimum_height=self.student_list.setter('height'))
        scroll.add_widget(self.student_list)
        list_box.add_widget(scroll)
        main.add_widget(list_box)
        
        # Grade panel
        grade_box = BoxLayout(orientation='vertical', size_hint_x=0.6)
        grade_box.add_widget(Label(text='Grade Entry', size_hint_y=None, height=dp(30), bold=True))
        
        self.info_label = Label(text='Select student', size_hint_y=None, height=dp(40))
        grade_box.add_widget(self.info_label)
        
        form = GridLayout(cols=2, spacing=dp(8), size_hint_y=0.6, padding=dp(10))
        
        form.add_widget(Label(text='Note (0-20):'))
        self.note_input = TextInput(multiline=False, input_filter='float')
        form.add_widget(self.note_input)
        
        form.add_widget(Label(text='Absent:'))
        self.absent_check = CheckBox()
        form.add_widget(self.absent_check)
        
        form.add_widget(Label(text='Justifié:'))
        self.justifie_check = CheckBox()
        form.add_widget(self.justifie_check)
        
        form.add_widget(Label(text='Note:'))
        self.obs_input = TextInput(multiline=True)
        form.add_widget(self.obs_input)
        
        grade_box.add_widget(form)
        
        btn_row = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(5))
        save_btn = ModernButton(text='Save Grade', color=SUCCESS_COLOR)
        save_btn.bind(on_press=self.save_grade)
        btn_row.add_widget(save_btn)
        
        view_btn = ModernButton(text='View All')
        view_btn.bind(on_press=self.view_all_grades)
        btn_row.add_widget(view_btn)
        grade_box.add_widget(btn_row)
        
        main.add_widget(grade_box)
        layout.add_widget(main)
        
        self.add_widget(layout)
        self.refresh_groups()
    
    def refresh_groups(self):
        groups = self.db.get_all_groups()
        for g in self.rosters.keys():
            if g not in groups:
                groups.append(g)
        groups.sort(key=lambda x: int(x.replace('Groupe', '')) if x.startswith('Groupe') else 999)
        self.group_spinner.values = groups
        if groups:
            self.group_spinner.text = groups[0]
    
    def on_group_selected(self, spinner, text):
        if text != 'Select':
            self.selected_groupe = text
            self.load_students()
    
    def on_section_change(self, spinner, text):
        if text.startswith('Section '):
            self.current_section = int(text.split()[1])
            if self.selected_student:
                self.load_grade()
    
    def load_roster(self, instance):
        if not self.selected_groupe or self.selected_groupe not in self.rosters:
            show_popup("No roster for this group")
            return
        
        roster = self.rosters[self.selected_groupe]
        success, errors, _ = self.db.import_students_from_roster(self.selected_groupe, roster)
        show_popup(f"Loaded {success} students\n{errors} duplicates skipped")
        self.load_students()
    
    def load_students(self):
        if not self.selected_groupe:
            return
        
        students = self.db.get_all_students(self.selected_groupe)
        self.student_list.clear_widgets()
        
        for s in students:
            btn = Button(
                text=f"{s['nom']}\n{s['prenom']}",
                size_hint_y=None,
                height=dp(60)
            )
            btn.bind(on_press=lambda x, st=s: self.select_student(st))
            self.student_list.add_widget(btn)
    
    def select_student(self, student):
        self.selected_student = student
        self.info_label.text = f"{student['nom']} {student['prenom']}"
        self.load_grade()
    
    def load_grade(self):
        if not self.selected_student:
            return
        
        grades = self.db.get_student_grades(self.selected_student['id'])
        gdict = {g['section_number']: g for g in grades}
        
        if self.current_section in gdict:
            g = gdict[self.current_section]
            self.note_input.text = str(g['note']) if g['note'] else ''
            self.absent_check.active = bool(g['absent'])
            self.justifie_check.active = bool(g['justifie'])
            self.obs_input.text = g['observation'] or ''
        else:
            self.note_input.text = ''
            self.absent_check.active = False
            self.justifie_check.active = False
            self.obs_input.text = ''
    
    def save_grade(self, instance):
        if not self.selected_student:
            show_popup("Select a student first")
            return
        
        note = None
        if self.note_input.text.strip():
            try:
                note = float(self.note_input.text.strip())
                if note < 0 or note > 20:
                    show_popup("Note must be 0-20")
                    return
            except:
                show_popup("Invalid note")
                return
        
        success, msg = self.db.add_or_update_grade(
            self.selected_student['id'],
            self.current_section,
            note,
            1 if self.absent_check.active else 0,
            1 if self.justifie_check.active else 0,
            self.obs_input.text.strip()
        )
        show_popup(msg, 'Success' if success else 'Error')
    
    def view_all_grades(self, instance):
        if not self.selected_student:
            show_popup("Select a student first")
            return
        
        grades = self.db.get_student_grades(self.selected_student['id'])
        avg, cnt = self.db.calculate_average(self.selected_student['id'])
        
        gdict = {g['section_number']: g for g in grades}
        text = f"{self.selected_student['nom']} {self.selected_student['prenom']}\n\n"
        text += f"Average: {avg if avg else 'N/A'} ({cnt} grades)\n\n"
        
        for i in range(1, Config.SEMESTER_SECTIONS + 1):
            if i in gdict:
                g = gdict[i]
                note_str = str(g['note']) if g['note'] else 'N/A'
                text += f"Section {i}: {note_str}\n"
            else:
                text += f"Section {i}: Not graded\n"
        
        show_popup(text, 'All Grades')
    
    def add_student(self, instance):
        if not self.selected_groupe:
            show_popup("Select a group first")
            return
        
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        
        form = GridLayout(cols=2, spacing=dp(8), size_hint_y=0.7)
        form.add_widget(Label(text='Matricule:'))
        mat_input = TextInput(multiline=False)
        form.add_widget(mat_input)
        
        form.add_widget(Label(text='Nom:'))
        nom_input = TextInput(multiline=False)
        form.add_widget(nom_input)
        
        form.add_widget(Label(text='Prénom:'))
        pre_input = TextInput(multiline=False)
        form.add_widget(pre_input)
        
        content.add_widget(form)
        
        btn_box = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(5))
        
        def do_add(x):
            if not all([mat_input.text, nom_input.text, pre_input.text]):
                show_popup("All fields required")
                return
            success, msg, _ = self.db.add_student(
                mat_input.text.strip(),
                nom_input.text.strip(),
                pre_input.text.strip(),
                self.selected_groupe
            )
            popup.dismiss()
            show_popup(msg)
            if success:
                self.load_students()
        
        add_btn = ModernButton(text='Add', color=SUCCESS_COLOR)
        add_btn.bind(on_press=do_add)
        btn_box.add_widget(add_btn)
        
        cancel_btn = ModernButton(text='Cancel', color=ERROR_COLOR)
        cancel_btn.bind(on_press=lambda x: popup.dismiss())
        btn_box.add_widget(cancel_btn)
        
        content.add_widget(btn_box)
        popup = Popup(title='Add Student', content=content, size_hint=(0.8, 0.6))
        popup.open()
    
    def export_grades(self, instance):
        if not self.selected_groupe:
            show_popup("Select a group first")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"grades_{self.selected_groupe}_{timestamp}.xlsx"
        
        if platform == 'android':
            path = os.path.join(get_storage_path(), 'StudentTracker')
        else:
            path = 'exports'
        
        os.makedirs(path, exist_ok=True)
        output = os.path.join(path, filename)
        
        success, msg = self.db.export_to_excel(output, self.selected_groupe)
        show_popup(f"{msg}\n\n{output}", 'Export')

# ============================================
# APP
# ============================================
class StudentTrackerApp(App):
    def build(self):
        self.title = Config.APP_NAME
        
        if platform == 'android':
            Clock.schedule_once(lambda dt: request_android_permissions(), 0.5)
        
        self.db = StudentTrackerDB()
        Window.clearcolor = BACKGROUND_COLOR
        
        sm = ScreenManager()
        sm.add_widget(MainScreen(name='main', db=self.db))
        
        logger.info("App started")
        return sm
    
    def on_stop(self):
        self.db.backup_database()

if __name__ == '__main__':
    StudentTrackerApp().run()
